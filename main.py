from summarizer import Summarizer
from summarizer.sbert import SBertSummarizer
from keybert import KeyBERT
# from summarizer.text_processors.coreference_handler import CoreferenceHandler
from transformers import pipeline


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

import re
import nltk
import time
from nltk import word_tokenize, pos_tag, ne_chunk
from nltk.corpus import mac_morpho
from nltk.tag import UnigramTagger
from nltk.tag import BigramTagger
from nltk.corpus import stopwords
import pandas as pd
import requests
from bs4 import BeautifulSoup
import heapq
from collections import Counter
from transformers import BertTokenizer, BertModel
import torch


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# nltk.download('punkt')
# nltk.download('punkt_tab')
# nltk.download('averaged_perceptron_tagger')
# nltk.download('mac_morpho')
# nltk.download('maxent_ne_chunker_tab')
# nltk.download('maxent_ne_chunker')
# nltk.download('words')
# nltk.download('stopwords')


def configurar_chromedriver():
    """ Função para configurar o ChromeDriver

    Returns:
        _type_: _description_
    """    
    service = Service("C:/Users/carlo/AppData/Local/Programs/Python/Python312/chromedriver.exe")
    chrome_options = Options()
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--headless")
    return service, chrome_options


def extrair_texto_selenium(url):
    """ Função para extrair texto de uma URL usando Selenium e chromedriver

    Args:
        url (string): link para notícia

    Returns:
        _type_: _description_
    """    
    service, chrome_options = configurar_chromedriver()

    driver = webdriver.Chrome(service=service, options=chrome_options)

    driver.get(url)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, 'span'))
        )
    except TimeoutException:
        print("A página não carregou a tempo.")
        driver.quit()
        return None
    
    html = driver.page_source
    driver.quit()
    
    soup = BeautifulSoup(html, 'html.parser')
    text = ''

    for paragraph in soup.find_all('span'):
        text += paragraph.get_text() + ' '
    
    text = re.sub(r'\s+', ' ', text)
    text = text.strip().lower()
    
    return text


def ler_links_arquivo(file_name):
    """ Função para ler os links de um arquivo Excel

    Args:
        file_name (_type_): _description_

    Returns:
        _type_: _description_
    """    
    df = pd.read_excel(file_name, engine='openpyxl')
    return df.iloc[0:, 1].tolist()


def extrair_conteudo_links(links):
    """ Função para extrair conteúdo dos links

    Args:
        links ([string]): _description_

    Returns:
        _type_: _description_
    """    
    textos = []
    for link in links:
        try:
            response = requests.get(link)
            response.raise_for_status()
            textos.append(response.text)
        except requests.exceptions.RequestException as e:
            print(f"Erro ao acessar {link}: {e}")
    return textos


def processar_texto_html(texto, link=None):
    """Função para processar e limpar o conteúdo HTML

    Args:
        texto (_type_): _description_
        link (_type_, optional): _description_. Defaults to None.

    Returns:
        _type_: _description_
    """    
    soup = BeautifulSoup(texto, 'html.parser')
    text = ''
    for paragraph in soup.find_all('p'):
        text += paragraph.get_text() + ' '

    text = re.sub(r'\s+', ' ', text)
    text = text.strip().lower()

    # se não houver texto extraído, tenta com Selenium
    if not text and link:
        text = extrair_texto_selenium(link)
    return text


def calcular_relevancia_texto(noticia):
    """ Função para calcular as sentenças mais relevantes com TF-IDF

    Args:
        noticia (string): noticia para ser analisada

    Returns:
        _type_: _description_
    """    
    vectorizer = TfidfVectorizer()

    sentences = nltk.sent_tokenize(noticia)

    # transforma o texto em vetores TF-IDF
    tfidf_matrix = vectorizer.fit_transform(sentences)

    # calcular a similaridade de cosseno entre as sentenças
    cosine_similarities = cosine_similarity(tfidf_matrix, tfidf_matrix)

    sentence_scores = cosine_similarities.mean(axis=1)

    # retorna 5 sentenças mais relevantes
    top_sentence_indices = heapq.nlargest(5, range(len(sentence_scores)), key=sentence_scores.__getitem__)

    return [sentences[i] for i in top_sentence_indices]


def processar_noticias_e_calcular_relevancia(file_name, modelo):
    """ Função principal para processar todas as notícias e calcular sentenças relevantes

    Args:
        file_name (string): _description_
    """    
    links = ler_links_arquivo(file_name)
    # textos = extrair_conteudo_links(links)

    # noticias = []
    # for index, texto in enumerate(textos):
    #     text = processar_texto_html(texto, link=links[index])
    #     noticias.append(text)
    #     print(text)
    #     print("\n" + "-" * 100 + "\n")

    # print("\n" + "-" * 100 + "\n")

    with open('textos.txt', 'r') as arquivo:
        noticias = [linha.strip() for linha in arquivo.readlines()]

    resumos = []
    resumos_bert = []
    for noticia in noticias:
        sentencas_relevantes = calcular_relevancia_texto(noticia)
        resumo = ""
        for i, sentenca in enumerate(sentencas_relevantes):
            if i < len(sentencas_relevantes) - 1:
                resumo += sentenca + " "
            else:
                resumo += sentenca
        # print(resumo)
        resumos.append(resumo)
        # print("\n" + "-" * 100 + "\n")

        # trechos = extrair_trechos_entre_aspas(noticia)
        # if trechos:
        #     print("Trechos encontrados entre aspas:")
        #     for trecho in trechos:
        #         print(trecho)
        #     print("\n" + "-" * 100 + "\n")
        
        resumo_bert = modelo(noticia)
        # print("Resumo Bert")
        # print(resumo_bert)
        # print("\n" + "-" * 100 + "\n")
        resumos_bert.append(resumo_bert)


    return links, resumos, resumos_bert

def extrair_trechos_entre_aspas(texto):
    """Extrai trechos entre aspas duplas e simples de um texto.

    Args:
        texto (str): O texto de onde os trechos serão extraídos.

    Returns:
        list: Uma lista de trechos encontrados entre aspas.
    """
    padrao = r'(["\'])(.*?)\1'
    trechos = re.findall(padrao, texto)

    return [trecho[1] for trecho in trechos]

def treinar_tagger_portugues():
    treino = mac_morpho.tagged_sents()
    unigram_tagger = UnigramTagger(treino)
    bigram_tagger = BigramTagger(treino, backoff=unigram_tagger)
    return bigram_tagger

def extrair_palavras_chaves(texto, tagger, keyBert):
    tokens = word_tokenize(texto)
    pos_tags = tagger.tag(tokens)
    
    if pos_tags is None:
        print("Erro: pos_tags é None.")
        return []
    
    # carrega stopwords em português
    stop_words = list(stopwords.words('portuguese'))
    
    frases_nominais = []
    frase_atual = []

    for palavra, tag in pos_tags:
        # N: Substantivo comum  NPROP: Substantivo próprio  ADJ: adjetivo
        if tag in ('N', 'NPROP', 'ADJ') and palavra not in stop_words:
            frase_atual.append(palavra)
        else:
            if frase_atual:
                frases_nominais.append(' '.join(frase_atual))
                frase_atual = []

    if frase_atual:
        frases_nominais.append(' '.join(frase_atual))
    
    # extraindo palavras-chave usando contagem de frequência
    contagem = Counter(token for token in tokens if token not in stop_words)
    principais_palavras = contagem.most_common(10)
    
    # extraindo palavras-chave usando TF-IDF
    texto_unido = ' '.join(frases_nominais)
    vectorizer = TfidfVectorizer(stop_words=stop_words)
    X = vectorizer.fit_transform([texto_unido])

    tfidf_scores = pd.DataFrame(X.T.toarray(), index=vectorizer.get_feature_names_out(), columns=["score"])
    principais_tfidf = tfidf_scores.nlargest(10, "score")




    palavras_chaves_bert = keyBert.extract_keywords(texto)

    return frases_nominais, principais_palavras, principais_tfidf.index.tolist(), palavras_chaves_bert

# MODELOS
# model = Summarizer() # Bert 301.840493 segundos 309.030349 segundos

# handler = CoreferenceHandler(greedyness=.4)
# model = Summarizer(sentence_handler=handler)

# model = Summarizer('distilbert-base-uncased', hidden=[-1,-2], hidden_concat=True) # Distil Bert 96.214338 segundos 86.871495 segundos

model = SBertSummarizer('paraphrase-MiniLM-L6-v2') # Sentence Bert 60.28 segundos 57.753437 segundos

kw_model = KeyBERT()


start_time = time.time()

file_name = 'docs_rafael.xlsx'

links, resumos, resumos_bert = processar_noticias_e_calcular_relevancia(file_name, model)

tagger = treinar_tagger_portugues()

todos_principais_tfidf = []
for resumo in resumos:
    frases_nominais, principais_palavras, palavras_chave_tf_idf, palavras_chave_bert = extrair_palavras_chaves(resumo, tagger, kw_model)
    todos_principais_tfidf.append(palavras_chave_tf_idf)

todos_principais_bert = []
for resumo in resumos_bert:
    frases_nominais, principais_palavras, palavras_chave_tf_idf, palavras_chave_bert = extrair_palavras_chaves(resumo, tagger, kw_model)
    todos_principais_bert.append(palavras_chave_bert)

vetor_ajustado = []
for vetor in todos_principais_bert:
    novo_vetor = []
    for item in vetor:
        novo_vetor.append(item[0])

    string_unica = ", ".join(novo_vetor)
    vetor_ajustado.append(string_unica)

end_time = time.time()

checkpoint = "Sami92/XLM-R-Large-ClaimDetection"
tokenizer_kwargs = {'padding':True,'truncation':True,'max_length':512}
claimdetection = pipeline("text-classification", model = checkpoint, tokenizer =checkpoint, **tokenizer_kwargs, device="cuda")

conjuncao_pattern = r'\b(e|nem|também|bem como|não só...mas|também|mas|porém|contudo|todavia|entretanto|no entanto|não obstante|ou|ou...ou|já...já|ora...ora|quer...quer|seja...seja|logo|pois|portanto|assim|por isso|por consequência|por conseguinte|porque|que|porquanto|visto que|uma vez que|já que|pois que|como|tanto que|tão que|tal que|tamanho que|de forma que|de modo que|de sorte que|de tal forma que|a fim de que|para que|quando|enquanto|agora que|logo que|desde que|assim que|tanto que|apenas|se|caso|desde|salvo se|exceto se|contando que|embora|conquanto|ainda que|mesmo que|se bem que|posto que|assim como|tal|qual|tanto como|conforme|consoante|segundo|à proporção que|à medida que|ao passo que|quanto mais...mais)\b|\.'

termos_relevantes_geral = []
for resumo in resumos_bert:
    partes = re.split(conjuncao_pattern, resumo)
    partes = [parte.strip() for parte in partes if parte is not None and parte.strip()]
    results = claimdetection(partes)
    termos_relevantes_resumo = []
    for i, result in enumerate(results):
        if result['label'] == 'factual':
            termos_relevantes_resumo.append(partes[i])
    termos_relevantes_geral.append(str(termos_relevantes_resumo))


df_resumos = pd.DataFrame({
    'Link': links,
    'Resumo': resumos,
    'Palavras-chave TF-IDF': [', '.join(t) for t in todos_principais_tfidf],
    'Resumos Bert': resumos_bert,
    'Palavras-chaves Bert': vetor_ajustado,
    'termos_relevantes': termos_relevantes_geral
})

df_resumos.to_excel('resultados/check_worthy_results.xlsx', index=False, engine='openpyxl')

wb = load_workbook('resultados/check_worthy_results.xlsx')
ws = wb.active

ws.column_dimensions[get_column_letter(1)].width = 20
ws.column_dimensions[get_column_letter(2)].width = 80
ws.column_dimensions[get_column_letter(3)].width = 25
ws.column_dimensions[get_column_letter(4)].width = 80
ws.column_dimensions[get_column_letter(5)].width = 25
ws.column_dimensions[get_column_letter(6)].width = 80

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

wb.save('resultados/check_worthy_results.xlsx')

print(f"\nTEMPO TOTAL: {end_time - start_time:.6f} segundos")
